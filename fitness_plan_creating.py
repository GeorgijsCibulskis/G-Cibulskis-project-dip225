from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time

def start_driver():
    service = Service()

    # Šī ir ļoti ērta funkcija, lai katru reizi, kad programma ir palaista, neatvērtos internets, bet tas strādātu fonā (lietotājs neredzēs, kā atveras tīmekļa vietne un kas tajā notiek)
    options = Options()
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=service, options=options)
    return driver

def consent(driver):
    consent_button = driver.find_element(By.CLASS_NAME, 'fc-cta-consent')
    consent_button.click()

# Funkcija, kura izvēlas lietotāja dzimumu un pārbauda uzreiz ievades kļūdas
def select_gender(driver):
    while True:
        gender = input("Are you male or female? ")
        if gender.lower() == 'male':

            # Selenium meklē elementu "Input", kura vērtība ir "male" vai "female" (tīmekļa vietnes kodā tas elements "input" ir klases "radio", kas principā ir poga, uz kuru var uzklikšķināt)
            driver.find_element(By.CSS_SELECTOR, 'input[value="male"]').click()
            break
        elif gender.lower() == 'female':
            driver.find_element(By.CSS_SELECTOR, 'input[value="female"]').click()
            break
        else:
            print("Invalid gender input. Please choose 'male' or 'female'.")
            continue
    return gender

# Funkcija, kura izvēlas ēdienreižu skaitu un pārbauda, vai tas tika pareizi ievadīts
def select_meal_count(driver):
    while True:
        meal_count = input("Enter how many times do you eat during the day (3-6): ")
        if (meal_count.isnumeric() == False) or int(meal_count) < 3 or int(meal_count) > 6:
            print("Invalid meals count. Please enter valid number.")
            continue
        else:

            # Tā kā tīmekļa lapa ir pārāk gara, un programma nevarēs redzēt atbilstošo pogu, mums ir nepieciešams pārvietot lapu uz leju līdz tam brīdim, kad būs redzama atbilstoša poga
            meal_count_element = driver.find_element(By.CSS_SELECTOR, f'input[value="{meal_count}"]')
            driver.execute_script("arguments[0].scrollIntoView();", meal_count_element)
            meal_count_element.click()
            break
    return meal_count

def get_info(driver, age, weight, goal_weight, height, time_span, exercise_mode, meal_count, gender):

    # Meklējām atbilstošus logus, kur ierakstīt nepieciešamu informāciju par lietotāju 
    driver.find_element(By.ID, 'fin_age').send_keys(age)
    driver.find_element(By.ID, 'fin_height_cm').send_keys(height)
    driver.find_element(By.ID, 'fin_curr_weight_kg').send_keys(weight)
    driver.find_element(By.ID, 'fin_goal_weight_kg').send_keys(goal_weight)

    # Objekts, kura klase ir Select (šī klase ir domāta tieši darbam ar tā sauktajiem dropdown elementiem tīmekļa vietnēs, kuri sastāv no vairākam izvēles variantiem)
    time_span_dropdown = Select(driver.find_element(By.ID, 'fin_goal_span'))

    # Select klases objektam ir metode select_by_visible_text, kura izvēlas kādu variantu un visiem iespējamiem pēc teksta, kurš ir rakstīts tajā
    time_span_dropdown.select_by_visible_text(time_span)

    
    exercise_dropdown = Select(driver.find_element(By.ID, 'fin_activity_modifier'))
    exercise_dropdown.select_by_visible_text(exercise_mode)

    # Atgriežam ievadīto informāciju tālākām darbībām
    return [gender.capitalize(), age, height, weight, goal_weight, time_span, exercise_mode, meal_count]

def get_calculated_info(driver):
    driver.find_element(By.CLASS_NAME, 'btn-danger').click()
    time.sleep(2)

    # Atrodam atbilstošu tabulu, tad atrodam tajā visas rindas, paņemam tikai 2. (ar indeksu 1) rindu un no 2 - 4 šūnas iegēstam barības vielu daudzumu dienā
    rows = driver.find_elements(By.XPATH, "//table[@id='tableContent']/tbody/tr")
    row_for_per_day = rows[1]

    # Šeit tiek paņemta 2., 3. un 4. pozīcija no tabulas 2 rindas, kuras atbilst par Carbs, Protein un Fat vērtībām dienā
    needed_cells = row_for_per_day.find_elements(By.XPATH, './td[position() > 1 and position() < 5]')
    nutrients_per_day = [round(float(nutrients.text)) for nutrients in needed_cells]

    # Kopējam informāciju no garas tabulas, kur ir pierakstīti kaloriju daudzumi katrā nedēļā
    rows = driver.find_elements(By.XPATH, "//table[@id='weight-gain-table']/tbody/tr")
    calories_data = []
    week_number = []
    for row in rows:
        cells = row.find_elements(By.TAG_NAME, "td")
        row_data = [cell.text.strip() for cell in cells]
        calories_data.append(round(float(row_data[1])))
        week_number.append(int(row_data[0]))

    return [nutrients_per_day, calories_data, week_number]

def correct_width(sheet, row):

    # Šeit ir cikls, kurš izmanto funkciju enumerate, kura atgriež tuple struktūras vērtības sekojošā veidā (numurs, vērtība), mūsu gadījumā numurs ir kolonnas numurs un vērtība ir
    # kāda tipa vērtība, kura tika ierakstīta excelī
    for col_num, value in enumerate(row, start=1):

        # Iegūstam cik garš ir vērtības teksts
        max_length = len(str(value))

        # Iegūstam kolonnas burtu no kolonnas numura
        col_letter = get_column_letter(col_num)

        # Izmainām atbilstošas kolonnas platumu (+3 ir lietots, lai teksts pilnībā būtu attēlots pareizā šūnā)
        sheet.column_dimensions[col_letter].width = max_length + 3

def merge_and_center(sheet, section_name, start_row, end_row, start_col, end_col):

    # Šī funkcija ir domāta, lai apvienotu vairākas šūnas kopā un centrēt tekstu tajā (principā merge and center poga pašā Excel, bet automatizētā un smukāk izskatās :) )
    merged_cell = sheet.cell(row=start_row, column=start_col, value=section_name)

    # Alignment klase ir daļa no openpyxl styles, kas ļauj modificēt šūnas saturu (izvietojums)
    merged_cell.alignment = Alignment(horizontal='center')

    # Šeit funkcija merge_cells, apvieno vairākas šūnas, balstoties uz uzdotām 'koordinatēm'
    sheet.merge_cells(start_row=merged_cell.row, start_column=merged_cell.column, end_row=end_row, end_column=end_col)

# Funkcija, kura uztaisa šūnām izvēlētajā diapazonā izceltas robežas
def border_for_sheets(sheet, start_row, end_row, start_col, max_col):

    # Robežu apraksts
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Cikls, kurš iet cauri rindām pa kolonnām un maina katrai šūnai robežas
    for row in sheet.iter_rows(min_row = start_row, max_row = end_row, min_col = start_col, max_col = max_col):
        for cell in row:
            cell.border = border_style

# Funkcija paņem rindu darba lapā un no tās katru aizpildītu šūnu, tad to centrē
def center_whole_row(sheet, row):
    for cell in sheet[row]:
        cell.alignment = Alignment(horizontal='center')

# Funkcija, kura pieraksta visu iegūto informāciju excelī
def write_to_excel(user_info, nutrients_per_day, nutrients_per_meal, week_data, calories_per_day, calories_per_meal, name):
    workbook = Workbook()
    sheet = workbook.active

    user_info_labels = ['Gender', 'Age', 'Height', 'Weight', 'Goal Weight', 'Time Span', 'Exercise Mode', 'Meal Count']
    sheet.append([name])
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = 2)
    sheet.append([''])
    sheet.append(user_info_labels)
    center_whole_row(sheet, sheet.max_row)
    sheet.append(user_info)
    center_whole_row(sheet, sheet.max_row)
    correct_width(sheet, user_info_labels)
    sheet.append([''])

    # Barības vielas visas dienas garumā
    nutrients_labels = ['Carbs per Day', 'Protein per Day', 'Fat per Day']

    # Uztaisa apvienotu šūnu vienā rindā un tik garu, cik garš ir barības vielu saraksts
    merge_and_center(sheet, 'Nutrients per day', sheet.max_row + 1, sheet.max_row + 1, 1, len(nutrients_labels))
    sheet.append(nutrients_labels)
    center_whole_row(sheet, sheet.max_row)
    sheet.append(nutrients_per_day)
    center_whole_row(sheet, sheet.max_row)
    correct_width(sheet, nutrients_labels)
    sheet.append(['']) 
    
    # Barības vielas vienā ēdienreizē
    nutrients_labels = ['Carbs per Meal', 'Protein per Meal', 'Fat per Meal']
    merge_and_center(sheet, 'Nutrients per meal', sheet.max_row + 1, sheet.max_row + 1, 1, len(nutrients_labels))
    sheet.append(nutrients_labels)
    center_whole_row(sheet, sheet.max_row)
    sheet.append(nutrients_per_meal)
    center_whole_row(sheet, sheet.max_row)
    correct_width(sheet, nutrients_labels)
    sheet.append([''])
    sheet.append([''])

    # Kaloriju daudzums katrā nedēļa (gan dienas, gan katras ēdienreizes 'deva')
    sheet.append(['Week Number', 'Calories per Day', 'Calories per Meal'])
    sheet.column_dimensions['C'].width = len('Calories per Meal') + 2
    center_whole_row(sheet, sheet.max_row)
    for week, calories_day, calories_meal in zip(week_data, calories_per_day, calories_per_meal):
        sheet.append([week, calories_day, calories_meal])
        center_whole_row(sheet, sheet.max_row)

    # Šie cipari būs vajadzīgi tālāk programma, lai kontrolētu kurā vietā excelī veikt izmaiņas
    sheet['D15'] = 'Control Column'
    sheet['E15'] = 'Control Day'
    sheet['F15'] = 'Control Row'
    sheet['G15'] = 'Control Week'
    sheet['D16'] = 1
    sheet['E16'] = 1
    sheet['F16'] = sheet.max_row + 3
    sheet['G16'] = 1

    # Uztaisa robežas visām tabulām un šūnām
    border_for_sheets(sheet, 1, 1, 1, 2)
    border_for_sheets(sheet, 3, 4, 1, 8)
    border_for_sheets(sheet, 6, 9, 1, 3)
    border_for_sheets(sheet, 10, 12, 1, 3)
    border_for_sheets(sheet, 15, sheet.max_row, 1, 3)
    
    sheet.column_dimensions['D'].width = len('Control Column') + 2
    center_whole_row(sheet, 15)
    center_whole_row(sheet, 16)

    current_date = datetime.now().date()
    sheet_name = f"Start - {current_date} ({user_info[5]})"
    sheet.title = sheet_name

    name = name.split(" ")
    name[0] = name[0][0]
    name = name[::-1]
    name = "_".join(name)
    workbook.save(f"{name}_nutrition.xlsx")

# Galvenā funkcija, kura palaiž visu failu TIKAI, ja pats fails tika izsaukts, nevis importēts kādā citā failā
def main(age, weight, goal_weight, height, time_span, exercise_mode, meal_count, gender, name, progressbar):

    driver = start_driver()
    driver.get("https://www.prokerala.com/health/health-calculators/weight-gain-calculator.php")
    time.sleep(2)

    consent(driver)
    info_about_user = get_info(driver, age, weight, goal_weight, height, time_span, exercise_mode, meal_count, gender)
    time.sleep(2)
    calculated_info = get_calculated_info(driver)
    nutrients_per_day = calculated_info[0]
    calories_per_day = calculated_info[1]
    calories_per_meal = [round(calories/int(info_about_user[7])) for calories in calories_per_day]
    week_data = calculated_info[2]
    nutrients_per_meal = [round(nutrient/int(info_about_user[7])) for nutrient in nutrients_per_day]

    write_to_excel(info_about_user, nutrients_per_day, nutrients_per_meal, week_data, calories_per_day, calories_per_meal, name)

    driver.close()
    # Lai apstādinātu progresa rādītāju, kad process ir beidzies (mēģināju vairākas metodes kā to izdarīt vienā failā user_interface.py, bet visas bija pārāk sarežģītas un izmantoja vēl vienu plūsmu, tāpēc
    # vieglāk ir vienkārši nosūtīt to uz main() kā argumentu)
    progressbar.pack_forget()

if __name__ == "__main__":
    main()

