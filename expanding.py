import time
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from fitness_plan_creating import merge_and_center, border_for_sheets
from searching import start_driver, consent
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def setup_excel(name, progressbar):
    try:
    # Funkcija, kura atver excel datni (lai katrā nākamā funkcija nerakstītu šo pašu, ir ērti to izdarīt vienu reizi un katrā nākamā funkcija kā argumentus sūtīt sheet, workbook un filename)
        name = name.split(" ")
        name[0] = name[0][0]
        name = name[::-1]
        name = "_".join(name)
        filename = f'{name}_nutrition.xlsx'
        workbook = load_workbook(filename = filename)
        sheet = workbook.active
        return workbook, sheet, filename
    
    except FileNotFoundError:
        progressbar.pack_forget()
        messagebox.showerror("Error", "No such file found - please enter valid name or create Your own fitness plan.")
        return None, None, None

# Funkcijas, kuras ir ērti izmantot, kad vajag ierakstīt excelī kaut ko rindā vai kolonnā, nevis lietot sheet.append, kas ieraksta nākamajā tukšā rindā, kas ne vienmēr ir vajadzīgs vai piemērots veids
def excel_column_append(sheet, start_row, start_column, appending_list):
    for index, value in enumerate(appending_list):
        existing_value = sheet.cell(row=start_row + index, column=start_column).value
        
        # Funkcija pārbauda vai šūnā jau ir kaut kas ierakstīts, ja jā, tad tam pieskaita argumenta vērtību (tas būs vajadzīgs tālāk, kad vienai un tai pašai dienai lietotājs gribēs pievienot vairākus produktus)
        if existing_value is not None:
            value = existing_value + value
        else:
            value = value
        
        # Funkcija pievieno atbilstošā šūnā vērtību
        sheet.cell(row=start_row + index, column=start_column, value=value)

def excel_row_append(sheet, start_row, start_column, appending_list):
    for index, value in enumerate(appending_list):
        sheet.cell(row = start_row, column = start_column + index, value = value)

# Nedaudz cita centrēšanas funkcija, kurā var uzdot KURA rinda ir jācentrē UN LĪDZ KURAI kolonnai
def center_whole_row(sheet, workbook, filename, row, start_col, end_col):

    # Iezīmējam rindu no excel (ar iter_rows var 'izcelt' tas šūnas, kuras ir vajadzīgas, jo tas funkcijas argumenti principā ir robežas no un līdz kuram ir jāiezīmē šūnas iekšā excelī)
    # Šī funkcija ņem visas šūnas (no min_col līdz max_col) no 1. rindas, tad visas no 2. rindas utt. līdz pēdējai rindai (max_row)
    for row in sheet.iter_rows(min_row = row, max_row = row, min_col=start_col, max_col=end_col):

        # Katru rindas šūnu pārtaisam par centrēto
        for cell in row:
            cell.alignment = Alignment(horizontal = 'center')
    workbook.save(filename)
    
# Funkcija, kura taisa smuko tabulu katrai nedēļai, kurā tiks attēloti katras dienas apēstas barības vielas un kalorijas
def expanding_week_excel(workbook, sheet, filename):

    # Uztaisa tabulas katrai šūnai robežu ('border')
    border_for_sheets(sheet, sheet['F16'].value, sheet['F16'].value + 5, 1, 8)

    day_numbers = [1, 2, 3, 4, 5, 6, 7]
    nutrients = ['Day', 'Carbs', 'Protein', 'Fat', 'Calories']
    merge_and_center(sheet, f"Week {sheet['G16'].value}", sheet['F16'].value, sheet['F16'].value, 2, 8)
    excel_row_append(sheet, sheet['F16'].value + 1, 2, day_numbers)
    excel_column_append(sheet, sheet['F16'].value + 1, sheet['D16'].value, nutrients)
    sheet['D16'] = sheet['D16'].value + 1
    workbook.save(filename)


def changing_days_and_weeks(sheet):

    # Ja mums ir beidzies 7. diena, tātad sākas jauna nedēļa un jāsamaina vērtības uz sākumvērtībām, takai pamainot 'control row', lai taisītu jauno tabulu zemāk
    if sheet["E16"].value == 7:
        sheet["D16"] = 1
        sheet["E16"] = 1
        sheet["F16"] = sheet["F16"].value + 9
        sheet["G16"] = sheet["G16"].value + 1
        return
    
    # Ja mums vienkārši beidzas diena, tad mainām tikai 'control column' un 'control day'
    sheet["D16"] = sheet["D16"].value + 1
    sheet["E16"] = sheet["E16"].value + 1

def appending_nutrition(workbook, sheet, filename, driver, product, mass, end_of_day):

    # Ja mums ir 1. kolonna, tad tas nozīme, ka sākas jauna nedēļa, tāpēc jātaisa jauna tabula un jācentre visa tālāka tabula
    if sheet["D16"].value == 1:
        expanding_week_excel(workbook, sheet, filename)
        for i in range(6):
            center_whole_row(sheet, workbook, filename, sheet["F16"].value + i, 1, 8)
    
    # Izvēlētā produkta meklēšana
    driver.find_element(By.CSS_SELECTOR, "a[onclick='toggleSearchBox();return false;']").click()
    driver.find_element(By.ID, 'search-box').send_keys(product)
    driver.find_element(By.ID, 'search-box').send_keys(Keys.ENTER)
    time.sleep(2)

    # Izradījās, ka var gadīties tā, ka, vēlreiz ievadot produkta nosaukumu (jau no piedāvāta saraksta, kurš atrodas teskta kastē) ievades laukā tīmekļa vietnē, tā vietne var atgriezt nevis uzreiz lapu ar barības vielām
    # tam produktam, bet atkal tabulu, kurā būs līdzīgie elementi (šo es uzzināju, kā produktu ievadot 'pork' un tad izvēloties 'Salami, pork, Italian, tad, ja meklēt to pēdējo, vietne atgriež tīmekļa lapu
    # vēl ar 3 produktiem). Šī iemesla dēļ vajag pārbaudīt, vai lapā nav elements ar tāda produkta nosaukumu, ja ir, tad uz to ir jāuzklikšķina
    if driver.find_elements(By.CSS_SELECTOR, f'a[title="{product}"]'):
        product_link = driver.find_element(By.CSS_SELECTOR, f'a[title="{product}"]')
        driver.execute_script("arguments[0].scrollIntoView();", product_link)
        product_link.click()
        time.sleep(2)

    # Ar kalorijām viss ir ērti - vietnē tiem ir atsevišķa ID
    calories = float(driver.find_element(By.ID, 'calories').text)

    # Ar taukiem un ogļhidrātiem jau ir sarežģītāk - vajadzēja atrast tabulas rindu, tad tajā <b> elementu, un tajā tekstu, bet vissarežģītākais bija atrast kā attēlot kodā elementu &nbsp;, lai selenium
    # varētu atrast pareizo ceļu, jo citādāk python radīja kļūdas
    fat_full_text = driver.find_element(By.XPATH, "//table[@class='center zero']/tbody/tr/td[@class='left'][b[text()='Total\u00a0Fat']]").text
    fat = (fat_full_text.split(" "))[2][:-1]
    fat = float(fat)
    carbs_full_text = driver.find_element(By.XPATH, "//table[@class='center zero']/tbody/tr/td[@class='left'][b[text()='Total\u00a0Carbohydrate']]").text
    carbs = (carbs_full_text.split(" "))[2][:-1]
    carbs = float(carbs)

    # Jauna sintakse kā ātrāk atrast olbaltumvielu daudzumu - nevis meklēt sākumā <b> elementu un tad no tā 'parent' lauku, bet uzreiz meklēt <td> elementu, kurā iekšā ir <b> elements,
    # kura iekšā ir teksts 'Protein'
    protein_td_element = driver.find_element(By.XPATH, "//table[@class='center zero']/tbody/tr/td[@class='left'][b[text()='Protein']]")
    protein = protein_td_element.text.split(" ")[1][:-1]
    protein = float(protein)

    # Atrodam kādai produkta masai ir atbilstošas barības vielas, lai izmantotu šo skaitli koeficienta rēķināšanai
    portion_size = float(driver.find_element(By.ID, "serving-size").text.split(" ")[0])

    coeficient = float(mass)/portion_size
    nutrients = [round(carbs*coeficient), round(protein*coeficient), round(fat*coeficient), round(calories*coeficient)]
    excel_column_append(sheet, sheet['F16'].value + 2, sheet['D16'].value, nutrients)

    # Pārbaudām, vai diena nav beidzies, ja ir, tad jāsamaina 'control day' un 'control column' vērtības
    if end_of_day == "Yes":
        changing_days_and_weeks(sheet)
    workbook.save(filename)
    return

def main(name, product, mass, end_of_day, progressbar):
    workbook, sheet, filename = setup_excel(name, progressbar)
    if workbook != None and sheet != None and filename != None:
        driver = start_driver()
        driver.get("https://www.nutritionvalue.org")
        time.sleep(2)

        consent(driver)
        
        try:
            appending_nutrition(workbook, sheet, filename, driver, product, mass, end_of_day)
        except PermissionError:
            messagebox.showerror("Permission Error", "Failed to save the workbook. Please make sure the file is not open.")
        progressbar.pack_forget()
        driver.close()

if __name__ == "__main__":
    main()