from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from fitness_plan_creating import merge_and_center
from fitness_plan_creating import center_whole_row
import time

def start_driver():
    service = Service()
    # Šī ir ļoti ērta funkcija, lai katru reizi, kad programma ir palaista, neatvertos internets, bet tas strādātu fonā (lietotājs neredzēs, kā atveras tīmekļa vietne un kas tajā notiek)
    options = Options()
    options.add_argument('--headless')
    driver = webdriver.Chrome(service=service, options=options)
    time.sleep(1)
    return driver

def setup_excel():
    # Funkcija, kura atver excel datni (lai katrā nākamā funkcija nerakstītu šo pašu, ir ērti to izdarīt vienu reizi un katrā nākamā funkcija kā argumentus sūtīt sheet un workbook)
    name = input("Please enter your name and surname: ")
    name = name.split(" ")
    name[0] = name[0][0]
    name = name[::-1]
    name = "_".join(name)
    filename = f'{name}_nutrition.xlsx'
    workbook = load_workbook(filename=filename)
    sheet = workbook.active
    return workbook, sheet, filename

def consent(driver):
    driver.find_element(By.CSS_SELECTOR, 'input[value="Accept"]').click()


def search_for_product(driver):
    # Lietotājs ievada kādu produktu un tad programma to ievada meklēšana laukā
    while True:
        product_name = input("Please enter product's name: ")
        driver.find_element(By.CSS_SELECTOR, "a[onclick='toggleSearchBox();return false;']").click()
        driver.find_element(By.ID, 'search-box').send_keys(product_name)
        driver.find_element(By.ID, 'search-box').send_keys(Keys.ENTER)
        # Pārbaudam vai produkts eksistē vietnē
        if validate_search(driver):
            print("Please enter valid product's name (no such product found in databse)!")
            continue
        break

# Šī funkcija pārbauda vai tīmekļa lapā ir paragrafs, ar tekstu 'No matching foods found.'. Ja tāds ir, tad tas produkts, kuru ievādīja lietotājs ir nepareizs vai neeksistē vietnē
def validate_search(driver):
    try:
        check = driver.find_element(By.XPATH, "//p[text()='No matching foods found.']")
        return True

    except NoSuchElementException:
        return False

# Funkcijas, kuras ir ērti izmantot, kad vajag ierakstīt excelī kaut ko rindā vai kolonnā, nevis lietot sheet.append, kas ieraksta nākamajā tukšā rindā, kas ne vienmēr ir vajadzīgs vai piemerots veids
def excel_column_append(sheet, start_row, start_column, appending_list):
    for index, value in enumerate(appending_list):
        existing_value = sheet.cell(row=start_row + index, column=start_column).value
        
        # Funkcija pārbauda vai šūnā jau ir kaut kas ierakstīts, ja jā, tad tam pieskaita argumenta vērtību (tas būs vajadzīgs tālāk, kad vienai un tai pašai dienai lietotājs gribēs peivienot vairākus produktus)
        if existing_value is not None:
            value = existing_value + value
        else:
            value = value
        
        # Funkcija pievieno atbilstošā šūnā vērtību
        sheet.cell(row=start_row + index, column=start_column, value=value)

def excel_row_append(sheet, start_row, start_column, appending_list):
    for index, value in enumerate(appending_list):
        sheet.cell(row = start_row, column = start_column + index, value = value)

# Funkcija, kura taisa smūko tabulu katrai nedēļai, kurā tiks attēloti katras dienas apēstas barības vielas un kalorijas
def expanding_excel(workbook, sheet, filename):

    day_numbers = [1, 2, 3, 4, 5, 6, 7]
    nutrients = ['Carbs', 'Protein', 'Fat', 'Calories']
    merge_and_center(sheet, f"Week {sheet['G16'].value}", sheet['F16'].value, sheet['F16'].value, 2, 8)
    excel_row_append(sheet, sheet['F16'].value + 1, 2, day_numbers)
    excel_column_append(sheet, sheet['F16'].value + 2, sheet['D16'].value, nutrients)
    # sheet['D16'] = sheet['D16'].value + 1 Do not needed now, since not completed whole algorithm
    workbook.save(filename)

def choice_list(workbook, sheet, filename, driver):
    # Šajā vietā ērti lietot CSS_SELECTOR XPATH vietā, jo katrai šūnai ir klase 'left'
    # Tātad programma atrod visas tabulas rindas un nokopē no tiem tekstu, kas principā ir garš saraksts ar ēdienu nosaukumiem, kuri tika atrasti pēc lietotāja ievādīta vārda
    cells = driver.find_elements(By.CSS_SELECTOR, "td.left")
    text_from_cells = [cell.text for cell in cells]
    text_from_cells = text_from_cells[1::]
    # Izvadam visus produktus
    print(f"Here is the list of food, which is connected with your choice: ")
    for food in text_from_cells:
        print(food)
    while True:
        # Lietotājs var izvēlēties no saraksta kādu konkrēto ēdienu
        next_step = input("Choose from the list (if not in list enter 'The end'): " ).capitalize()
        if next_step == 'The end':
            return
        elif next_step not in text_from_cells:
            print("Product not in list, please enter correctly: ")
            continue
        else:
            # Jā tāds produkts ir sarakstā, tad uzspipiežam uz to
            true_product = driver.find_element(By.CSS_SELECTOR, f'a[title="{next_step}"]')
            driver.execute_script("arguments[0].scrollIntoView();", true_product)
            true_product.click()
            time.sleep(2)

            # Ar kalorijām viss ir ērti - vietnē tiem ir atsevišķa ID
            calories = float(driver.find_element(By.ID, 'calories').text)

            # Ar taukiem un ogļhidrātiem jau ir sarežģītāk - vajadzēja atrast tabulu un tad konkrēto pēc kārtas rindu un no tās dabūt tekstu un to apstrādāt līdz float vērtībai
            fat_full_text = driver.find_element(By.XPATH, "//table[@class='center zero']/tbody/tr[10]").text
            fat = (fat_full_text.split(" "))[2][:-1]
            fat = float(fat)
            carbs_full_text = driver.find_element(By.XPATH, "//table[@class='center zero']/tbody/tr[18]").text
            carbs = (carbs_full_text.split(" "))[2][:-1]
            carbs = float(carbs)

            # Jauna sintakse kā ātrāk atrast olbaltumvielu daudzumu - nevis meklēt sākumā <b> elementu un tad no tā 'parent' lauku, bet uzreiz meklēt <td> elementu, kurā iekšā ir <b> elements,
            # kura iekšā ir teksts 'Protein'
            protein_td_element = driver.find_element(By.XPATH, "//td[b[text()='Protein']]")
            protein = protein_td_element.text.split(" ")[1][:-1]
            protein = float(protein)

            # Atrodam kādai produkta masai ir atbilstošas barības vielas, lai izmanototu šo skaitli koeficienta rēķināšanai
            portion_size = int(driver.find_element(By.ID, "serving-size").text.split(" ")[0])

            mass = int(input("Please enter approximate mass of this product, that you have eaten (in gramms): "))
            coeficient = mass/portion_size
            nutrients = [round(carbs*coeficient), round(protein*coeficient), round(fat*coeficient), round(calories*coeficient)]
            excel_column_append(sheet, sheet['F16'].value + 2, sheet['D16'].value, nutrients)
            
            # Cikls, kurš centrē kolonnas, kurās tika ierakstīti dati
            for i in range(5):
                center_whole_row(sheet, sheet['F16'].value + i + 2)
            workbook.save(filename)
            return

def main():
    driver = start_driver()
    workbook, sheet, filename = setup_excel()
    driver.get("https://www.nutritionvalue.org")
    time.sleep(2)

    consent(driver)

    search_for_product(driver)
    choice_list(workbook, sheet, filename, driver)
    # expanding_excel(workbook, sheet, filename)
    # search_for_product()
    # product_calories = choice_list()
    # driver.close()

if __name__ == "__main__":
    main()